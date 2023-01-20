from openpyxl import *

wb = load_workbook('planilha.xlsx')
ws = wb.active

vendedor_1 = ws['A2']
vendedor_2 = ws['A3']
vendedor_3 = ws['A4']
vendedor_4 = ws['A5']

filiacao_1 = ws['B2']
filiacao_2 = ws['B3']
filiacao_3 = ws['B4']
filiacao_4 = ws['B5']

total_de_vendas_1 = ws['C2']
total_de_vendas_2 = ws['C3']
total_de_vendas_3 = ws['C4']
total_de_vendas_4 = ws['C5'].value

# print([vendedor_1.value], filiacao_1.value, [total_de_vendas_1.value])
# print([vendedor_2.value], filiacao_2.value, [total_de_vendas_2.value])
# print([vendedor_3.value], filiacao_3.value, [total_de_vendas_3.value])
# print([vendedor_4.value], filiacao_4.value, [total_de_vendas_4])

print(col_range)

# nova_celula = ws['C6']
#
# tot_vendas = 0
# for i in range(2, 6):
#     tot_vendas += int(ws[f'C{i}'].value)
#
# ws['C6'] = tot_vendas
#
